# plans/views.py  — drop-in замена

from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods, require_GET
from django.conf import settings
from django.utils import timezone
from django.core.cache import cache
from django.contrib.auth.decorators import login_required
from .models import Plan  # <-- ВАЖНО: импорт модели
from decimal import Decimal
from django.views.decorators.csrf import csrf_exempt


import openpyxl, csv
import requests
from datetime import datetime, timedelta

# -------------------- ХЕЛПЕРЫ ДАТ --------------------

def _iso_monday(d=None):
    """Понедельник (локальной недели) для даты d (или сегодня)."""
    d = timezone.localdate() if d is None else d
    return d - timedelta(days=d.weekday())

def _week_bounds(week_index: int):
    """
    1 = текущая неделя, 2 = предыдущая, 3 = -2, 4 = -3.
    Возвращает (start_date, end_date_exclusive).
    """
    start = _iso_monday() - timedelta(days=7 * (week_index - 1))
    end = start + timedelta(days=7)  # правая граница эксклюзивная
    return start, end

def _parse_dt(s: str):
    """Разбирает WB даты (saleDate/lastChangeDate) в datetime."""
    if not s:
        return None
    s = s.replace('Z', '+00:00')
    try:
        return datetime.fromisoformat(s)
    except Exception:
        try:
            return datetime.fromisoformat(s.split('.')[0])
        except Exception:
            return None

def _parse_date_ymd(s: str):
    """YYYY-MM-DD -> date | None"""
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except Exception:
        return None

# -------------------- ВЬЮХИ СТРАНИЦЫ --------------------

def bdds_page(request):
    """Главная страница с таблицей BDDS."""
    plans = Plan.objects.all().order_by('sku')
    return render(request, 'BDDS.html', {'plans': plans})

def dashboard(request):
    """Простая страница приветствия с быстрыми ссылками."""
    user_name = request.user.get_short_name() or request.user.username
    quick_links = [
        {"title": "План/Факт (БДДС)", "url_name": "bdds_page",   "icon": "bar-chart-2"},
        {"title": "Загрузить план",    "url_name": "bdds_upload", "icon": "upload"},
        {"title": "Главная",           "url_name": "index",       "icon": "home"},
    ]
    return render(request, "dashboard.html", {
    "user_name": user_name,
    "quick_links": quick_links,
})

@login_required
def unit_page(request):
    # Дефолтные профили (правь под себя)
    profiles = [
        {
            "id": "cosmetics",
            "name": "Косметика",
            "wb_rate": 0.25,   # комиссия WB
            "tax_rate": 0.04,  # налог
            "acos": 0.12,      # рекламные расходы
            "logistics": 65.0,
            "storage": 3.0,
            "returns": 5.0,
        },
        {
            "id": "clothing",
            "name": "Одежда",
            "wb_rate": 0.25,
            "tax_rate": 0.04,
            "acos": 0.10,
            "logistics": 120.0,
            "storage": 7.0,
            "returns": 20.0,
        },
    ]
    return render(request, "unit.html", {"profiles": profiles})
# -------------------- API: загрузка Excel --------------------

@require_http_methods(["POST"])
def upload_plans(request):
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'ok': False, 'error': 'Файл не получен'}, status=400)

    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        headers = {
            (str(ws.cell(row=1, column=c).value).strip().lower()
             if ws.cell(row=1, column=c).value is not None else ''): c
            for c in range(1, ws.max_column + 1)
        }
        if 'sku' not in headers or 'planperweek' not in headers:
            return JsonResponse({'ok': False, 'error': 'В файле должны быть столбцы SKU и PlanPerWeek'}, status=400)

        for r in range(2, ws.max_row + 1):
            sku_val = ws.cell(row=r, column=headers['sku']).value
            plan_val = ws.cell(row=r, column=headers['planperweek']).value
            if sku_val is None:
                continue
            sku = str(sku_val).strip()
            try:
                plan = int(plan_val) if plan_val is not None else 0
            except Exception:
                plan = 0

            Plan.objects.update_or_create(
                sku=sku,
                defaults={'plan_per_week': max(0, plan)}
            )

        data = list(Plan.objects.all().order_by('sku').values('sku', 'plan_per_week'))
        return JsonResponse({'ok': True, 'data': data})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

# -------------------- API: планы --------------------

def api_plans(request):
    data = list(Plan.objects.all().order_by('sku').values('sku', 'plan_per_week'))
    return JsonResponse({'ok': True, 'data': data})

# -------------------- API: WB факт (период ИЛИ неделя) --------------------

@require_GET
def wb_facts(request):
    """
    GET /bdds/api/facts/?week=1
    ЛИБО
    GET /bdds/api/facts/?start=YYYY-MM-DD&end=YYYY-MM-DD

    Считает факт-продажи по всем SKU за выбранную неделю или произвольный период.
    Матчит и по supplierArticle (строка), и по nmId (число).
    """
    token = getattr(settings, 'WB_API_KEY', '')
    if not token:
        return JsonResponse({'ok': False, 'error': 'WB_API_KEY пуст в settings/.env'}, status=500)

    # Период имеет приоритет над неделей
    s = request.GET.get('start')
    e = request.GET.get('end')
    start_date = _parse_date_ymd(s) if s else None
    end_date_incl = _parse_date_ymd(e) if e else None

    if start_date and end_date_incl and start_date <= end_date_incl:
        start = start_date
        end = end_date_incl + timedelta(days=1)  # эксклюзивная правая граница
        week = 0
    else:
        try:
            week = int(request.GET.get('week', '1'))
            if week not in (1, 2, 3, 4):
                week = 1
        except Exception:
            week = 1
        start, end = _week_bounds(week)

    # SKU из БД
    ordered = list(Plan.objects.order_by('sku').values_list('sku', flat=True))
    vendor_map = {str(s).strip().lower(): str(s).strip() for s in ordered if str(s).strip()}
    nmids = set()
    for ssku in ordered:
        s_str = str(ssku).strip()
        if s_str.isdigit():
            try:
                nmids.add(int(s_str))
            except Exception:
                pass

    # Агрегатор результата
    agg = {str(s).strip(): {'total': 0, 'days': [0] * 7} for s in ordered}

    # Запрос к WB
    headers = {'Authorization': token}
    params = {'dateFrom': start.strftime('%Y-%m-%dT00:00:00'), 'flag': 0}
    url = getattr(settings, 'WB_API_SALES_URL', 'https://statistics-api.wildberries.ru/api/v1/supplier/sales')

    try:
        resp = requests.get(url, headers=headers, params=params, timeout=getattr(settings, 'WB_API_TIMEOUT', 30))
        resp.raise_for_status()
        records = resp.json()
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'WB API error: {e}'}, status=502)

    # Агрегируем
    for r in records:
        art = str(r.get('supplierArticle') or r.get('article') or '').strip().lower()
        nm = r.get('nmId') or r.get('nm') or r.get('nmID')

        dt_str = r.get('saleDate') or r.get('date') or r.get('lastChangeDate')
        dt = _parse_dt(dt_str)
        if not dt:
            continue
        d = timezone.localtime(dt).date() if timezone.is_aware(dt) else dt.date()
        if not (start <= d < end):
            continue

        key = vendor_map.get(art)
        if key is None:
            try:
                nm_int = int(nm) if nm is not None else None
            except Exception:
                nm_int = None
            if nm_int is not None and nm_int in nmids:
                key = str(nm_int)

        if key is None:
            continue

        idx = d.weekday()  # 0..6 (Пн..Вс)
        agg[key]['total'] += 1
        agg[key]['days'][idx] += 1

    data = [{'sku': s, **agg[str(s).strip()]} for s in ordered]
    # если был произвольный период — вернём end как включительную дату
    end_out = (end - timedelta(days=1)).isoformat() if week == 0 else end.isoformat()
    return JsonResponse({'ok': True, 'week': week, 'start': start.isoformat(), 'end': end_out, 'data': data})

# -------------------- (опционально) отладка WB --------------------

@require_GET
def wb_peek(request):
    """
    GET /bdds/api/wb/peek/?dateFrom=YYYY-MM-DD
    Возвращает несколько сырых записей WB (для проверки сопоставления).
    """
    token = getattr(settings, 'WB_API_KEY', '')
    if not token:
        return JsonResponse({'ok': False, 'error': 'WB_API_KEY пуст'}, status=500)

    date_from = request.GET.get('dateFrom')
    if not date_from:
        date_from = _iso_monday().strftime('%Y-%m-%d')

    headers = {'Authorization': token}
    params = {'dateFrom': f'{date_from}T00:00:00', 'flag': 0}
    url = getattr(settings, 'WB_API_SALES_URL', 'https://statistics-api.wildberries.ru/api/v1/supplier/sales')

    try:
        resp = requests.get(url, headers=headers, params=params, timeout=getattr(settings, 'WB_API_TIMEOUT', 30))
        resp.raise_for_status()
        records = resp.json()
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'WB API error: {e}'}, status=502)

    sample = []
    for r in records[:25]:
        sample.append({
            'saleDate': r.get('saleDate') or r.get('date') or r.get('lastChangeDate'),
            'supplierArticle': r.get('supplierArticle') or r.get('article'),
            'nmId': r.get('nmId') or r.get('nm') or r.get('nmID'),
            'barcode': r.get('barcode'),
            'quantity': r.get('quantity') or 1
        })
    return JsonResponse({'ok': True, 'count': len(records), 'dateFrom': date_from, 'sample': sample})

def _parse_date_ymd(s: str):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except Exception:
        return None

@require_GET
def wb_facts_bucketed(request):
    """
    GET /bdds/api/facts/bucketed/
      [опц] ?start=YYYY-MM-DD&end=YYYY-MM-DD  -> фильтруем по периоду
    Возвращает факты сразу для недель 1..4:
      {
        ok: true,
        weeks: {
          "1": { start, end, days_in_period, data: [{sku,total,days:[7]}...] },
          "2": {...}, "3": {...}, "4": {...}
        }
      }
    Каждая продажа WB мапится в НУЖНУЮ неделю и НУЖНЫЙ день (Пн=0..Вс=6)
    ещё на сервере.
    """
    token = getattr(settings, 'WB_API_KEY', '')
    if not token:
        return JsonResponse({'ok': False, 'error': 'WB_API_KEY пуст'}, status=500)

    # 4 недели на экране
    weeks_bounds = {i: _week_bounds(i) for i in (1, 2, 3, 4)}
    # общий диапазон, который точно покрывает 4 недели
    start0 = min(s for (s, _) in weeks_bounds.values())
    end0   = max(e for (_, e) in weeks_bounds.values())

    # необязательный период
    s = request.GET.get('start')
    e = request.GET.get('end')
    start_filter = _parse_date_ymd(s) if s else None
    end_filter_incl = _parse_date_ymd(e) if e else None
    if start_filter and end_filter_incl and start_filter <= end_filter_incl:
        # расширим общий диапазон, чтобы точно захватить период
        start0 = min(start0, start_filter)
        end0   = max(end0, end_filter_incl + timedelta(days=1))  # эксклюзивная правая граница

    # кэш
    cache_key = f"wb_bucket:{start0.isoformat()}:{end0.isoformat()}:{start_filter or 'x'}:{end_filter_incl or 'x'}"
    cached = cache.get(cache_key)
    if cached:
        return JsonResponse(cached)

    # SKU
    ordered = list(Plan.objects.order_by('sku').values_list('sku', flat=True))
    vendor_map = {str(s).strip().lower(): str(s).strip() for s in ordered if str(s).strip()}
    nmids = set()
    for s in ordered:
        st = str(s).strip()
        if st.isdigit():
            try: nmids.add(int(st))
            except: pass

    # заготовка по неделям
    per_week = {
        i: {str(s).strip(): {'total': 0, 'days': [0]*7} for s in ordered}
        for i in (1,2,3,4)
    }

    # быстро определить к какой из 4 недель относится дата
    def week_index_of(d):
        for i,(ws,we) in weeks_bounds.items():
            if ws <= d < we:
                return i
        return None

    # считаем сколько ДНЕЙ каждой недели реально попадает в (необязательный) период
    def days_in_period(ws, we):
        if not (start_filter and end_filter_incl):
            return 7
        a = max(ws, start_filter)
        b = min(we, end_filter_incl + timedelta(days=1))
        if a >= b:
            return 0
        return (b - a).days

    weeks_days_in_period = {i: days_in_period(*weeks_bounds[i]) for i in (1,2,3,4)}

    # вызов WB один раз с нижней общей границы
    headers = {'Authorization': token}
    params = {'dateFrom': start0.strftime('%Y-%m-%dT00:00:00'), 'flag': 0}
    url = getattr(settings, 'WB_API_SALES_URL', 'https://statistics-api.wildberries.ru/api/v1/supplier/sales')

    try:
        import requests
        resp = requests.get(url, headers=headers, params=params, timeout=getattr(settings, 'WB_API_TIMEOUT', 30))
        resp.raise_for_status()
        records = resp.json()
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'WB API error: {e}'}, status=502)

    # разбор и сразу раскладка по НЕДЕЛЯМ и ДНЯМ
    for r in records:
        art = str(r.get('supplierArticle') or r.get('article') or '').strip().lower()
        nm  = r.get('nmId') or r.get('nm') or r.get('nmID')

        dt_str = r.get('saleDate') or r.get('date') or r.get('lastChangeDate')
        dt = _parse_dt(dt_str)
        if not dt:
            continue
        d = timezone.localtime(dt).date() if timezone.is_aware(dt) else dt.date()

        if not (start0 <= d < end0):
            continue
        # если задан период — ещё раз отфильтруем
        if start_filter and end_filter_incl:
            if not (start_filter <= d <= end_filter_incl):
                continue

        wk = week_index_of(d)
        if not wk:
            continue  # дата не попала ни в одну из 4 недель на экране

        # сопоставление sku
        key = vendor_map.get(art)
        if key is None:
            try:
                nm_int = int(nm) if nm is not None else None
            except Exception:
                nm_int = None
            if nm_int is not None and nm_int in nmids:
                key = str(nm_int)
        if key is None:
            continue

        idx = d.weekday()  # 0..6 (Пн..Вс)
        per_week[wk][key]['total'] += 1
        per_week[wk][key]['days'][idx] += 1

    # упакуем ответ
    weeks_payload = {}
    for i in (1,2,3,4):
        s, e = weeks_bounds[i]
        weeks_payload[str(i)] = {
            'start': s.isoformat(),
            'end': e.isoformat(),
            'days_in_period': weeks_days_in_period[i],
            'data': [{'sku': ssku, **per_week[i][str(ssku).strip()]} for ssku in ordered]
        }

    payload = {'ok': True, 'weeks': weeks_payload}
    cache.set(cache_key, payload, timeout=180)
    return JsonResponse(payload)

def index(request):
    return render(request, "index.html")


@login_required(login_url='login')
def dashboard(request):
    # Лёгкий контекст для приветствия и быстрых ссылок
    user_name = request.user.get_short_name() or request.user.username
    quick_links = [
        {"title": "План/Факт (БДДС)", "url_name": "bdds_page",   "icon": "bar-chart-2"},
        {"title": "Загрузить план",    "url_name": "bdds_upload", "icon": "upload"},
        {"title": "Главная",           "url_name": "index",       "icon": "home"},
    ]
    return render(request, "plans/dashboard.html", {
        "user_name": user_name,
        "quick_links": quick_links,
    })

















@login_required
def import_cost_form(request):
    """Импорт себестоимости через HTML-форму."""
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            # Если файл не был передан – показать страницу с сообщением об ошибке
            return render(request, 'plans/import_form.html', {'error': 'Файл не был загружен.'})
        try:
            # Определяем тип файла по расширению имени
            filename = file.name.lower()
            if filename.endswith('.csv'):
                # Обработка CSV-файла
                decoded_file = file.read().decode('utf-8', errors='ignore')
                reader = csv.reader(decoded_file.splitlines())
                headers = next(reader, None)
                if headers is None:
                    return render(request, 'plans/import_form.html', {'error': 'CSV-файл пустой.'})
                # Приводим заголовки к нижнему регистру для сопоставления
                headers = [h.strip().lower() for h in headers]
                if 'sku' not in headers or not ('cost' in headers or 'costprice' in headers):
                    return render(request, 'plans/import_form.html',
                                  {'error': 'В файле должны быть колонки "SKU" и "Cost" (или "CostPrice").'})
                sku_index = headers.index('sku')
                cost_index = headers.index('cost') if 'cost' in headers else headers.index('costprice')
                for row in reader:
                    if not row or len(row) <= max(sku_index, cost_index):
                        continue
                    sku_val = row[sku_index]
                    cost_val = row[cost_index] if cost_index < len(row) else ''
                    if not sku_val:
                        continue  # пустой артикул пропускаем
                    sku = str(sku_val).strip()
                    # Парсим себестоимость (заменяем запятую на точку, чтобы корректно привести к числу)
                    try:
                        cost = Decimal(str(cost_val).replace(',', '.')) if cost_val not in (None, '') else Decimal('0')
                    except Exception:
                        cost = Decimal('0')
                    # Обновляем или создаем товар
                    Plan.objects.update_or_create(
                        sku=sku,
                        defaults={'cost_price': cost}
                    )
            else:
                # Обработка Excel-файла (.xlsx или .xls)
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                # Считываем заголовки из первой строки
                headers = {
                    (str(ws.cell(row=1, column=c).value).strip().lower() if ws.cell(row=1, column=c).value else ''): c
                    for c in range(1, ws.max_column + 1)
                }
                if 'sku' not in headers or not ('cost' in headers or 'costprice' in headers):
                    return render(request, 'plans/import_form.html',
                                  {'error': 'В файле должны быть столбцы "SKU" и "Cost" (или "CostPrice").'})
                # Индексы колонок
                sku_col = headers['sku']
                cost_col = headers['costprice'] if 'costprice' in headers else headers['cost']
                # Проходим по строкам (со 2-й строки, т.к. 1-я – заголовки)
                for r in range(2, ws.max_row + 1):
                    sku_val = ws.cell(row=r, column=sku_col).value
                    cost_val = ws.cell(row=r, column=cost_col).value
                    if sku_val is None:
                        continue
                    sku = str(sku_val).strip()
                    # Парсим значение себестоимости из ячейки
                    if cost_val is None or str(cost_val).strip() == '':
                        cost = Decimal('0')
                    else:
                        try:
                            cost = Decimal(str(cost_val).replace(',', '.'))
                        except Exception:
                            cost = Decimal('0')
                    Plan.objects.update_or_create(
                        sku=sku,
                        defaults={'cost_price': cost}
                    )
            # Если импорт прошел успешно – отобразить страницу с сообщением об успехе
            return render(request, 'plans/import_form.html', {'success': True})
        except Exception as e:
            # В случае ошибки при чтении/парсинге файла – показать сообщение об ошибке
            return render(request, 'plans/import_form.html', {'error': f'Ошибка импорта: {e}'})
    else:
        # GET-запрос: отображаем страницу с формой для загрузки
        return render(request, 'plans/import_form.html')


@csrf_exempt
def import_cost_api(request):
    """Импорт себестоимости через API (принимает файл и возвращает JSON)."""
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Требуется POST-запрос.'}, status=400)
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'ok': False, 'error': 'Файл не найден в запросе.'}, status=400)
    try:
        filename = file.name.lower()
        created_count = 0
        updated_count = 0
        if filename.endswith('.csv'):
            # Аналогичная обработка CSV, как выше
            decoded_file = file.read().decode('utf-8', errors='ignore')
            reader = csv.reader(decoded_file.splitlines())
            headers = next(reader, None)
            if headers is None:
                return JsonResponse({'ok': False, 'error': 'CSV-файл пустой.'}, status=400)
            headers = [h.strip().lower() for h in headers]
            if 'sku' not in headers or not ('cost' in headers or 'costprice' in headers):
                return JsonResponse({'ok': False, 'error': 'Ожидаются колонки "SKU" и "Cost"/"CostPrice".'}, status=400)
            sku_index = headers.index('sku')
            cost_index = headers.index('cost') if 'cost' in headers else headers.index('costprice')
            for row in reader:
                if not row or len(row) <= max(sku_index, cost_index):
                    continue
                sku_val = row[sku_index]
                cost_val = row[cost_index] if cost_index < len(row) else ''
                if not sku_val:
                    continue
                sku = str(sku_val).strip()
                try:
                    cost = Decimal(str(cost_val).replace(',', '.')) if cost_val not in (None, '') else Decimal('0')
                except Exception:
                    cost = Decimal('0')
                obj, created = Plan.objects.update_or_create(
                    sku=sku,
                    defaults={'cost_price': cost}
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        else:
            # Обработка XLSX
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = {
                (str(ws.cell(row=1, column=c).value).strip().lower() if ws.cell(row=1, column=c).value else ''): c
                for c in range(1, ws.max_column + 1)
            }
            if 'sku' not in headers or not ('cost' in headers or 'costprice' in headers):
                return JsonResponse({'ok': False, 'error': 'Ожидаются столбцы "SKU" и "Cost"/"CostPrice".'}, status=400)
            sku_col = headers['sku']
            cost_col = headers['costprice'] if 'costprice' in headers else headers['cost']
            for r in range(2, ws.max_row + 1):
                sku_val = ws.cell(row=r, column=sku_col).value
                cost_val = ws.cell(row=r, column=cost_col).value
                if sku_val is None:
                    continue
                sku = str(sku_val).strip()
                if cost_val is None or str(cost_val).strip() == '':
                    cost = Decimal('0')
                else:
                    try:
                        cost = Decimal(str(cost_val).replace(',', '.'))
                    except Exception:
                        cost = Decimal('0')
                obj, created = Plan.objects.update_or_create(
                    sku=sku,
                    defaults={'cost_price': cost}
                )
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        # Формируем успешный ответ с информацией о количестве обновленных/созданных записей
        return JsonResponse({
            'ok': True,
            'updated': updated_count,
            'created': created_count
        })
    except Exception as e:
        return JsonResponse({'ok': False, 'error': f'Ошибка при обработке файла: {e}'}, status=500)